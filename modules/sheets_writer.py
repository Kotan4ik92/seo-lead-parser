"""
Красивый Excel-вывод результатов парсера.
Цветовое кодирование, фиксированные ширины, заморозка шапки.
"""

import csv
from datetime import datetime
from pathlib import Path

from modules.seo_scanner import SeoResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

import config

# ─── Цвета ───────────────────────────────────────────────
CLR = {
    "header":    "1F3864",   # тёмно-синий
    "critical":  "C00000",   # красный
    "hot":       "E26B0A",   # оранжевый
    "warm":      "F6C142",   # жёлтый
    "cold":      "4472C4",   # синий
    "dead":      "808080",   # серый
    "red_cell":  "FFCCCC",
    "yellow_cell":"FFF2CC",
    "green_cell": "E2EFDA",
    "alt_row":   "F2F2F2",
    "white":     "FFFFFF",
}

TEMP_COLOR = {
    "💥": CLR["critical"],
    "🔥": CLR["hot"],
    "🌤": CLR["warm"],
    "❄️": CLR["cold"],
    "⛔": CLR["dead"],
}

# ─── Структура колонок ───────────────────────────────────
# (заголовок, ширина)
COLUMNS = [
    ("№",               4),
    ("URL",             35),
    ("Запрос",          18),
    ("Лид",             12),   # температура
    ("Балл",             6),
    ("Contact Email",   28),
    ("Статус",          14),
    ("Title",           30),
    ("Title AI",        10),
    ("Meta Desc",       30),
    ("Meta AI",         10),
    ("H1",              25),
    ("H1 AI",           10),
    ("Schema AI",       10),
    ("Robots AI",       10),
    ("AI вердикт",      45),
    ("HTTPS",            7),
    ("Sitemap",          9),
    ("Robots.txt",       9),
    ("Lang",             6),
    ("Viewport",         8),
    ("Проблемы",        30),
    ("Примечания",      20),
]

HEADERS = [c[0] for c in COLUMNS]
WIDTHS  = [c[1] for c in COLUMNS]


def _score_cell(score: int) -> tuple[str, str]:
    """Возвращает (текст, цвет фона) для AI балла."""
    if score < 0:
        return "н/д", CLR["white"]
    if score <= 3:
        return f"🔴 {score}/10", CLR["red_cell"]
    if score <= 6:
        return f"🟡 {score}/10", CLR["yellow_cell"]
    return f"🟢 {score}/10", CLR["green_cell"]


def _yn(val: bool) -> str:
    return "✅" if val else "❌"


def _trunc(text: str, n: int) -> str:
    return (text[:n] + "…") if len(text) > n else text


def _row_data(i: int, seo: SeoResult, score: int,
              temp: str, query: str,
              contact_email: str = "", status: str = "") -> list:
    ai = seo.ai
    tech_issues = "; ".join(
        iss for iss in seo.issues if not iss.startswith("[AI]")
    ) or "—"

    return [
        i,
        seo.url,
        query,
        temp,
        score,
        contact_email or "—",
        status or "🟡 New",
        _trunc(seo.title, 60) or "—",
        ai.title_score,
        _trunc(seo.meta_desc, 60) or "—",
        ai.meta_score,
        _trunc(seo.h1_text, 50) or "—",
        ai.h1_score,
        ai.schema_score,
        ai.robots_score,
        _trunc(ai.ai_verdict, 120) if ai.ai_verdict else "—",
        _yn(seo.https),
        _yn(seo.has_sitemap),
        _yn(seo.has_robots),
        _yn(seo.lang_set),
        _yn(seo.viewport_set),
        tech_issues,
        "",   # Примечания
    ]


def write_to_excel(results: list[tuple[SeoResult, int, str]],
                   query: str, filename: str = "leads.xlsx",
                   contacts: dict = None, statuses: dict = None) -> None:
    if not XLSX_OK:
        print("[!] openpyxl не установлен — сохраняю в CSV")
        write_to_csv(results, query)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    # ── Шапка ──────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=CLR["header"])
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Данные ─────────────────────────────────────────
    for row_idx, (seo, score, temp) in enumerate(results, 2):
        email  = (contacts or {}).get(seo.url, "")
        status = (statuses or {}).get(seo.url, "")
        data = _row_data(row_idx - 1, seo, score, temp, query, email, status)
        alt = (row_idx % 2 == 0)

        # Цвет строки по температуре
        first_char = temp[:1] if temp else ""
        row_color = TEMP_COLOR.get(first_char, CLR["white"])

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)

            # Цветные AI-балл ячейки (колонки 7,9,11,12,13)
            if col_idx in (7, 9, 11, 12, 13) and isinstance(value, int):
                txt, bg = _score_cell(value)
                cell.value = txt
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            # Колонка температуры — цвет по лиду
            elif col_idx == 4:
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.font = Font(bold=True, size=9, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 5:  # Балл
                cell.alignment = Alignment(horizontal="center", vertical="top")
                if score >= 70:
                    cell.font = Font(bold=True, color="C00000", size=9)
                elif score >= 45:
                    cell.font = Font(bold=True, color="E26B0A", size=9)
            elif alt and col_idx not in (7, 9, 11, 12, 13, 4):
                cell.fill = PatternFill("solid", fgColor=CLR["alt_row"])

        ws.row_dimensions[row_idx].height = 40

    # ── Итог внизу ─────────────────────────────────────
    last = len(results) + 2
    hot  = sum(1 for r, s, t in results if s >= 45)
    warm = sum(1 for r, s, t in results if 20 <= s < 45)
    cold = sum(1 for r, s, t in results if s < 20 and r.reachable)
    ws.cell(row=last, column=1,
            value=f"Итого: {len(results)} сайтов  |  💥🔥 {hot} горячих  |  🌤 {warm} тёплых  |  ❄️ {cold} холодных"
            ).font = Font(bold=True, size=10)

    wb.save(filename)
    print(f"[EXCEL] Сохранено в {filename}")


def write_to_csv(results: list[tuple[SeoResult, int, str]],
                 query: str, filename: str = "leads.csv") -> None:
    today = datetime.now().strftime("%d.%m.%Y %H:%M")
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for i, (seo, score, temp) in enumerate(results, 1):
            writer.writerow(_row_data(i, seo, score, temp, query))
    print(f"[CSV] Сохранено в {filename}")
